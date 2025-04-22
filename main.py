from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import load_workbook, Workbook


url = "https://vkusnoitochka.ru/menu"
firefox_options = FirefoxOptions()
firefox_options.binary_location = r"C:\Program Files\Mozilla Firefox\firefox.exe"
firefox_options.set_preference("permissions.default.geo", 2) 
firefox_options.set_preference("geo.enabled", False)
firefox_options.add_argument("--headless")


service = FirefoxService(r"C:\web\geckodriver\geckodriver.exe")


driver = webdriver.Firefox(service=service, options=firefox_options)
wait = WebDriverWait(driver, 10)

def load_existing_data(file_path):
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        data = {}
        for row in ws.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовки
            name, price, image = row
            data[name] = {"price": price, "image": image}
        return data, wb, ws
    except FileNotFoundError:
        return {}, None, None


def save_data_to_excel(file_path, data, wb=None, ws=None):
    if not wb or not ws:
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(["Name", "Price", "Image"])  

    
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    
    for name, info in data.items():
        ws.append([name, info["price"], info["image"]])

    wb.save(file_path)

try:
    driver.get(url)
    print("Заголовок страницы:", driver.title)
    burgers_rolls_section = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#burgery-i-rolly")))
    burgers_rolls_section.click()
    time.sleep(2)  
    items = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".product-card")))
    file_path = "products.xlsx"
    existing_data, wb, ws = load_existing_data(file_path)
    for item in items:
        try:
            driver.execute_script("arguments[0].scrollIntoView(true);", item)
            time.sleep(1)  
            time.sleep(0.5)
            item.click()
            modal = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".modal-abstraction")))
            time.sleep(1)
            name_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".product-top__name")))
            name = name_element.text.strip()
            price = modal.find_element(By.CSS_SELECTOR, "span.font-type-6:nth-child(1)").text.strip()
            image = modal.find_element(By.CSS_SELECTOR, ".common-image__img").get_attribute("src")
            if name in existing_data:
                if existing_data[name]["price"] != price or existing_data[name]["image"] != image:
                    print(f"Обновлен товар: {name}")
                    existing_data[name] = {"price": price, "image": image}
            else:
                print(f"Добавлен новый товар: {name}")
                existing_data[name] = {"price": price, "image": image}

        except Exception as e:
            print(f"Ошибка при парсинге товара: {e}")

        finally:
            try:
                close_button = modal.find_element(By.CSS_SELECTOR, ".modal-abstraction__close")
                close_button.click()
                wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".modal-abstraction")))
            except Exception:
                pass

finally:
    save_data_to_excel(file_path, existing_data, wb, ws)
    driver.quit()